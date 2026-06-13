# -*- coding: utf-8 -*-
"""
承包方家庭成员 & 承包地块信息提取工具 v0.3.5
从 D:\农经全延保 下所有表1/表2 xlsx 文件中提取信息，汇总到一张表。
表1：同一户中同一人如在确权区和变动区都出现，合并为一条记录。减少的人口自动删除。
表2：提取地块确权登记信息及地块增减变动情况。
"""

import os
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ── 常量 ──────────────────────────────────────────────────────────────────────

OUTPUT_COLUMNS = (
    "所属组", "承包方编码", "承包方代表",
    "联系电话", "发包方名称", "户内人口",
    "家庭成员姓名", "性别", "身份证号", "身份证+性别核对",
    "与承包方代表关系", "变动情况",
    "分、合户来源", "调查记事(附记)",
)

OUTPUT_COLUMNS_B2 = (
    "所属组", "承包方编码", "承包合同编号", "承包方代表",
    "联系方式", "确权总面积(亩)",
    "地块总数", "地块序号",
    "地块名称", "地块编码", "地块面积(亩)",
    "东至", "西至", "南至", "北至",
    "变动情况", "变动面积(亩)", "变动原因",
    "分、合户来源", "调查记事",
)

_HOUSEHOLD_KEY_COLS = (0, 1)

_REDUCE_REMOVE_KEYWORDS = ("嫁", "死", "亡", "去世", "登记错误")

_FILL_A = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
_FILL_B = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_SEP_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_ERR_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# 身份证校验码权重和映射
_ID_WEIGHTS = (7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2)
_ID_CHECK_CODES = "10X98765432"


def _validate_id_checksum(id_number):
    """校验18位身份证号的第18位校验码，返回 True/False。"""
    if len(id_number) != 18 or not id_number[:17].isdigit():
        return False
    s = sum(int(id_number[i]) * _ID_WEIGHTS[i] for i in range(17))
    return id_number[17].upper() == _ID_CHECK_CODES[s % 11]

# 关系 → 预期性别（本人/配偶无法推断，不在此表）
_REL_MALE = frozenset(("孙子", "儿子", "父亲", "丈夫", "兄", "弟", "侄子", "外甥"))
_REL_FEMALE = frozenset(("孙女", "女儿", "母亲", "妻子", "姐", "妹", "侄女", "外甥女"))

# ── 户主变更时的关系转换规则 ─────────────────────────────────────────────────
# 键: 新户主在确权区的旧关系（相对于老户主）
# 值: 字典，映射 旧关系→新关系（相对于新户主）
# 无法准确推算的关系保持不变（如侄子、外甥等）
_RELATION_TRANSFORMS = {
    "儿子": {
        "本人": "父亲",   "妻子": "母亲",   "儿子": "兄弟",
        "女儿": "姐妹",   "父亲": "祖父",   "母亲": "祖母",
        "孙子": "儿子",   "孙女": "女儿",
    },
    "女儿": {
        "本人": "父亲",   "妻子": "母亲",   "儿子": "兄弟",
        "女儿": "姐妹",   "父亲": "祖父",   "母亲": "祖母",
        "孙子": "儿子",   "孙女": "女儿",
    },
    "妻子": {
        "本人": "丈夫",
    },
    "丈夫": {
        "本人": "妻子",
    },
    "孙子": {
        "本人": "祖父",   "妻子": "祖母",   "儿子": "父亲",
        "女儿": "母亲",
    },
    "孙女": {
        "本人": "祖父",   "妻子": "祖母",   "儿子": "父亲",
        "女儿": "母亲",
    },
    "父亲": {
        "本人": "儿子",   "妻子": "儿媳",   "儿子": "兄弟",
        "女儿": "姐妹",   "父亲": "本人",   "母亲": "母亲",
    },
    "母亲": {
        "本人": "儿子",   "妻子": "儿媳",   "儿子": "兄弟",
        "女儿": "姐妹",   "父亲": "父亲",   "母亲": "本人",
    },
}


def _recalculate_relationships(members, old_head, new_head, new_head_old_rel):
    """当户主变更时，重新计算所有家庭成员与新户主的关系。"""
    if not new_head_old_rel or new_head_old_rel == "本人":
        return  # 无需转换

    transform = _RELATION_TRANSFORMS.get(new_head_old_rel)
    if not transform:
        return  # 未知转换规则，跳过

    for m in members:
        name = m.get("家庭成员姓名", "").strip()
        old_rel = m.get("与承包方代表关系", "").strip()

        if name == new_head:
            m["与承包方代表关系"] = "本人"
        elif name == old_head:
            m["与承包方代表关系"] = transform.get("本人", "父亲")
        else:
            new_rel = transform.get(old_rel)
            if new_rel:
                m["与承包方代表关系"] = new_rel


def _check_gender(id_number, gender_text, relationship):
    """核对性别 + 身份证长度 + 校验码。返回 '正确'/'错误'/''。"""
    id_number = str(id_number or "").strip()
    gender_text = str(gender_text or "").strip()
    relationship = str(relationship or "").strip()

    if not gender_text:
        return ""

    has_id = bool(id_number) and id_number not in ("/", "／")

    # ① 身份证长度校验
    if has_id and len(id_number) != 18:
        return "错误"

    # ② 身份证校验码校验
    if has_id and not _validate_id_checksum(id_number):
        return "错误"

    # ③ 身份证性别核对
    id_gender = ""
    if has_id and id_number[:17].isdigit():
        id_gender = "男" if int(id_number[-2]) % 2 == 1 else "女"

    # ④ 关系核对（本人/配偶跳过）
    rel_gender = ""
    if relationship in _REL_MALE:
        rel_gender = "男"
    elif relationship in _REL_FEMALE:
        rel_gender = "女"

    # 两项都不确定 → 留空
    if not id_gender and not rel_gender:
        return ""

    # 任一项与表中性别不符 → 错误
    if id_gender and id_gender != gender_text:
        return "错误"
    if rel_gender and rel_gender != gender_text:
        return "错误"
    return "正确"


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def _is_biao1(filename):
    """判断是否为表1文件，兼容 表1.xlsx / 表1（1）.xlsx / 表1--总(分户).xlsx 等"""
    return bool(re.search(r'-表1[^\\]*\.xlsx$', filename)) and not filename.startswith("~$")


def _is_biao2(filename):
    """判断是否为表2文件，兼容 表2.xlsx / 表2（1）.xlsx 等"""
    return bool(re.search(r'-表2[^\\]*\.xlsx$', filename)) and not filename.startswith("~$")


def _normalize_biao_suffix(filename):
    """把 表1（1）.xlsx / 表1 副本.xlsx 等统一为 表1.xlsx，用于解析文件名"""
    return re.sub(r'(表[12])[^.]*\.xlsx$', r'\1.xlsx', filename)


def _parse_filename(filename):
    fn = _normalize_biao_suffix(filename)
    # 从原始文件名提取分户信息（归一化会去掉）
    fenhuyi = ""
    m_fh = re.search(r'[（(](分户[^)]*)[）)]', filename)
    if m_fh:
        fenhuyi = m_fh.group(1).strip()
    # 先匹配 编号-姓名-表1.xlsx
    m = re.match(r"^(.+?)-(.+?)-表1\.xlsx$", fn)
    if m:
        code, name = m.group(1), m.group(2)
        if fenhuyi:
            name = f"{name}（{fenhuyi}）"
        return code, name
    # 再匹配 编号-姓名表1.xlsx（只有一个横线）
    m = re.match(r"^(.+?)-(.+?)表1\.xlsx$", fn)
    if m:
        code, name = m.group(1), m.group(2)
        if fenhuyi:
            name = f"{name}（{fenhuyi}）"
        return code, name
    # 无横线时整体作为编号
    base = fn.replace("表1.xlsx", "")
    return base.rstrip("-"), ""


def _parse_filename2(filename):
    fn = _normalize_biao_suffix(filename)
    # 先匹配 编号-姓名-表2.xlsx
    m = re.match(r"^(.+?)-(.+?)-表2\.xlsx$", fn)
    if m:
        return m.group(1), m.group(2)
    # 再匹配 编号-姓名表2.xlsx（只有一个横线）
    m = re.match(r"^(.+?)-(.+?)表2\.xlsx$", fn)
    if m:
        return m.group(1), m.group(2)
    # 无横线时整体作为编号
    base = fn.replace("表2.xlsx", "")
    return base.rstrip("-"), ""


def _normalize(text):
    return str(text).replace(" ", "").replace("\n", "").replace("\r", "")


def _find_section_header(ws, keyword):
    # 使用更具体的关键词避免误匹配（如"有无确权承包合同"被误认为确权表头）
    specific = {"确权": "确权登记", "变动": "变动情况"}
    kw = specific.get(keyword, keyword)
    for r in range(1, ws.max_row + 1):
        # 只在C1列查找（表头行的标题都在C1）
        val = ws.cell(r, 1).value
        if val and kw in _normalize(val):
            return r
    return None


def _is_seq_number(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return val > 0
    s = str(val).strip()
    try:
        return float(s) > 0
    except (ValueError, TypeError):
        return False


# ── 表1 数据读取 ──────────────────────────────────────────────────────────────

def _read_section_data(ws, header_row, end_row=None, include_change_col=False):
    results = []
    max_r = (end_row - 1) if end_row else ws.max_row
    for r in range(header_row + 1, max_r + 1):
        seq = ws.cell(r, 2).value
        name = ws.cell(r, 4).value
        if not name or str(name).strip() == "":
            continue
        # 确权区必须有序号；变动区允许无序号的"子行"（续行）
        if not include_change_col and not _is_seq_number(seq):
            continue
        # 跳过分户区表头行（含有"承包方""代表关系""基础信息"等关键词）
        name_str = str(name).strip()
        if any(kw in name_str for kw in ("承包方", "代表关系", "基础信息", "成员总数")):
            continue
        # 跳过性别列明显不是性别值的行（如"□有""☑无"等）
        gender_val = str(ws.cell(r, 6).value or "").strip()
        if gender_val and gender_val not in ("男", "女", ""):
            if "□" in gender_val or "☑" in gender_val or "有" in gender_val or "无" in gender_val:
                continue
        row_data = {
            "家庭成员姓名": name_str,
            "性别": gender_val,
            "身份证号": str(ws.cell(r, 7).value or "").strip(),
            "与承包方代表关系": str(ws.cell(r, 9).value or "").strip(),
            "变动情况": str(ws.cell(r, 3).value or "").strip() if include_change_col else "",
        }
        if include_change_col:
            row_data["_reason"] = str(ws.cell(r, 13).value or "").strip()
        results.append(row_data)
    return results


def _person_key(row):
    id_num = row.get("身份证号", "").strip()
    if id_num and id_num != "/" and id_num != "／":
        return ("id", id_num)
    return ("name", row.get("家庭成员姓名", "").strip())


def _should_remove(row):
    if row.get("变动情况", "") != "减少":
        return False
    return True


# ── 表1 解析 ──────────────────────────────────────────────────────────────────

def parse_biao1(filepath, group_name):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    ws = wb.active

    fn = os.path.basename(filepath)
    file_code, file_name = _parse_filename(fn)

    contractor = str(ws.cell(4, 5).value or file_name or "").strip()
    phone = str(ws.cell(4, 10).value or "").strip()
    fa_bao_fang = str(ws.cell(3, 1).value or "").strip()
    fa_bao_fang = fa_bao_fang.replace("发包方名称：", "")
    fa_bao_fang = fa_bao_fang.replace("发包方名称:", "").strip()

    # 分户文件表头布局不同，读到的是表头文本而非数据 → 回退到文件名
    _header_kw = ("承包方", "基础信息", "代表关系", "□有", "☑无", "□是", "☑是", "成员总数")
    if any(kw in contractor for kw in _header_kw):
        contractor = file_name or ""
    if any(kw in fa_bao_fang for kw in _header_kw):
        fa_bao_fang = ""
    if any(kw in phone for kw in _header_kw):
        phone = ""

    # 提取承包方编码：确权承包合同编号去掉末尾字母
    contract_no = str(ws.cell(5, 9).value or "").strip()
    contractor_code = contract_no.rstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz") if contract_no else (file_code or "")

    info = {
        "所属组": group_name,
        "编号": file_code or "",
        "承包方编码": contractor_code,
        "承包方代表": contractor,
        "联系电话": phone,
        "发包方名称": fa_bao_fang,
    }

    h1 = _find_section_header(ws, "确权")
    h2 = _find_section_header(ws, "变动")

    section1 = _read_section_data(ws, h1, end_row=h2, include_change_col=False) if h1 else []
    section2 = _read_section_data(ws, h2, include_change_col=True) if h2 else []
    # 提取调查记事(附记)：扫描确权区和变动区所有行的第14列
    jishi_parts = []
    if h1:
        end1 = (h2 - 1) if h2 else ws.max_row
        for r in range(h1 + 1, end1 + 1):
            v = str(ws.cell(r, 14).value or "").strip()
            if v and "记事" not in v and v not in jishi_parts:
                jishi_parts.append(v)
    if h2:
        for r in range(h2 + 1, ws.max_row + 1):
            v = str(ws.cell(r, 14).value or "").strip()
            if v and "记事" not in v and v not in jishi_parts:
                jishi_parts.append(v)
    info["调查记事(附记)"] = "；".join(jishi_parts)
    wb.close()

    s1_key_idx = {}
    s1_name_idx = {}  # name -> index (all entries, for name fallback)
    for i, row in enumerate(section1):
        s1_key_idx[_person_key(row)] = i
        s1_name_idx[row.get("家庭成员姓名", "").strip()] = i

    # 记录确权区的老户主（关系为"本人"的人）
    old_head = ""
    for row in section1:
        if row.get("与承包方代表关系", "").strip() == "本人":
            old_head = row.get("家庭成员姓名", "")
            break

    s2_only = []
    _new_head_old_rel = ""  # 新户主在确权区的旧关系（更新前保存）
    for row in section2:
        pk = _person_key(row)
        # 优先用 person_key 匹配
        idx = s1_key_idx.get(pk)
        # 备用：person_key 未命中时，按姓名匹配（覆盖变动区补录身份证号的场景）
        if idx is None:
            name = row.get("家庭成员姓名", "").strip()
            if name in s1_name_idx:
                idx = s1_name_idx[name]
        if idx is not None:
            section1[idx]["变动情况"] = row["变动情况"]
            section1[idx]["_reason"] = row.get("_reason", "")
            # 确权区身份证号为空或"/"时，用变动区的值同步
            id_old = section1[idx].get("身份证号", "")
            id_new = row.get("身份证号", "")
            if id_new and id_new != "/" and (not id_old or id_old == "/"):
                section1[idx]["身份证号"] = id_new
                # 同步性别（变动区有真实身份证号时，性别更可靠）
                new_gender = row.get("性别", "").strip()
                if new_gender:
                    section1[idx]["性别"] = new_gender
            # 变动区姓名与确权区不同时（如错别字纠正），用变动区的值
            new_name = row.get("家庭成员姓名", "").strip()
            old_name = section1[idx].get("家庭成员姓名", "").strip()
            if new_name and new_name != old_name:
                section1[idx]["家庭成员姓名"] = new_name
            new_rel = row.get("与承包方代表关系", "").strip()
            if new_rel:
                # "户主" 等同于 "本人"，统一为 "本人"
                is_head = new_rel in ("本人", "户主")
                if is_head:
                    new_rel = "本人"
                # 在更新前，记录新户主的旧关系（用于后续重算）
                if is_head:
                    _new_head_old_rel = section1[idx].get("与承包方代表关系", "").strip()
                section1[idx]["与承包方代表关系"] = new_rel
                if is_head:
                    # 新户主出现，在变动情况后备注
                    head_new = row["家庭成员姓名"]
                    if old_head and head_new != old_head:
                        note = "更换户主：%s（原：%s）" % (head_new, old_head)
                        cur_chg = section1[idx].get("变动情况", "")
                        section1[idx]["变动情况"] = (cur_chg + "（" + note + "）") if cur_chg else note
        else:
            s2_only.append(row)

    # ── 户主变更时，重算所有家庭成员与新户主的关系 ──
    # 但如果变动区已有"变更关系"的明确标注，则以变动区为准，不再重算
    new_head_name = None
    for row in section1:
        if row.get("与承包方代表关系", "").strip() == "本人":
            new_head_name = row.get("家庭成员姓名", "").strip()
            break
    _has_explicit_changes = any(
        row.get("_reason", "").strip() == "变更关系" for row in section2
    )
    if (new_head_name and old_head and new_head_name != old_head
            and _new_head_old_rel and not _has_explicit_changes):
        _recalculate_relationships(section1, old_head, new_head_name, _new_head_old_rel)

    merged = section1 + s2_only

    # 解析分户：从变动原因中提取分户信息
    # 匹配 "分户为户主X" 或 "分户为户X的下面"
    split_groups = {}  # group_number -> [person_key, ...]
    person_reasons = {}  # person_key -> {group_number: reason}
    for row in section2:
        reason = row.get("_reason", "")
        if "分户" in reason:
            m = re.search(r"分户为户(?:主)?(\d+)", reason)
            if m:
                gn = int(m.group(1))
                pk = _person_key(row)
                split_groups.setdefault(gn, []).append(pk)
                person_reasons.setdefault(pk, {})[gn] = reason

    # 按分户分组：group 0 = 原户, 1+ = 分户
    # person_to_groups: person_key -> sorted list of group numbers
    person_to_groups = {}
    for gn, pks in split_groups.items():
        for pk in pks:
            person_to_groups.setdefault(pk, []).append(gn)
    for pk in person_to_groups:
        person_to_groups[pk].sort()

    groups = {}
    for row in merged:
        pk = _person_key(row)
        pgs = person_to_groups.get(pk)
        if pgs is None:
            # 不属于任何分户，归入原户
            groups.setdefault(0, []).append(row)
        else:
            # 将此人添加到其所属的所有分户组
            for i, gn in enumerate(pgs):
                pr = person_reasons.get(pk, {}).get(gn, "")
                if i == 0:
                    row["_group_reason"] = pr
                    groups.setdefault(gn, []).append(row)
                else:
                    cp = dict(row)
                    cp["_group_reason"] = pr
                    groups.setdefault(gn, []).append(cp)

    # 处理 _reason 和 _group_reason，追加到变动情况
    for gn, rows in groups.items():
        for row in rows:
            reason = row.pop("_reason", "")
            gr = row.pop("_group_reason", "")
            effective = gr if gr else reason
            if effective and effective != "无":
                cur = row.get("变动情况", "")
                row["变动情况"] = (cur + "（" + effective + "）") if cur else effective

    # 检查分户组是否缺少人员
    for gn, pks in split_groups.items():
        if gn not in groups:
            groups[gn] = []
        existing_pks = [_person_key(r) for r in groups[gn]]
        for pk in pks:
            if pk not in existing_pks:
                for row in merged:
                    if _person_key(row) == pk:
                        cp = dict(row)
                        pr = person_reasons.get(pk, {}).get(gn, "")
                        if pr and pr != "无":
                            cur = cp.get("变动情况", "")
                            cp["变动情况"] = (cur + "（" + pr + "）") if cur else pr
                        groups[gn].append(cp)
                        break

    if not groups:
        groups[0] = []

    result = []
    for gn in sorted(groups.keys()):
        if gn == 0:
            result.append((info, groups[gn]))
        else:
            gi = dict(info)
            gi["_split_group"] = gn
            gi["_split_source"] = info.get("编号", "")
            for r in groups[gn]:
                if r.get("与承包方代表关系", "").strip() == "本人":
                    gi["承包方代表"] = r["家庭成员姓名"]
                    break
            result.append((gi, groups[gn]))
    return result


# ── 表2 解析 ──────────────────────────────────────────────────────────────────

def parse_biao2(filepath, group_name):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    ws = wb.active

    fn = os.path.basename(filepath)
    file_code, file_name = _parse_filename2(fn)

    contractor = str(ws.cell(4, 4).value or file_name or "").strip()
    phone = str(ws.cell(4, 6).value or "").strip()
    total_area = str(ws.cell(4, 10).value or "").strip()
    total_plots = str(ws.cell(4, 14).value or "").strip()

    # 承包方编码：从文件名前缀提取
    contractor_code = file_code or ""
    contract_no = (file_code + "J") if file_code else ""

    info = {
        "所属组": group_name,
        "编号": file_code or "",
        "承包方编码": contractor_code,
        "承包合同编号": contract_no,
        "承包方代表": contractor,
        "联系方式": phone,
        "确权总面积(亩)": total_area,
        "地块总数": total_plots,
    }

    h1 = _find_section_header(ws, "确权")
    h2 = _find_section_header(ws, "变动")

    # 读取调查记事(确权区C13列，跳过表头行)
    jishi_parts = []
    if h1:
        end_j = (h2 - 1) if h2 else ws.max_row
        for r in range(h1 + 1, end_j + 1):
            v = str(ws.cell(r, 13).value or "").strip()
            if v and "记事" not in v and v not in jishi_parts:
                jishi_parts.append(v)
    info["调查记事"] = "；".join(jishi_parts)

    # Read confirmed plots
    plots = []
    if h1:
        end = (h2 - 1) if h2 else ws.max_row
        for r in range(h1 + 1, end + 1):
            name = ws.cell(r, 3).value
            if not name or str(name).strip() == "":
                continue
            plots.append({
                "地块名称": str(name).strip(),
                "地块编码": str(ws.cell(r, 4).value or "").strip(),
                "地块面积(亩)": str(ws.cell(r, 6).value or "").strip(),
                "东至": str(ws.cell(r, 7).value or "").strip(),
                "西至": str(ws.cell(r, 8).value or "").strip(),
                "南至": str(ws.cell(r, 9).value or "").strip(),
                "北至": str(ws.cell(r, 10).value or "").strip(),
            })

    # Read changes
    changes = {}
    if h2:
        for r in range(h2 + 1, ws.max_row + 1):
            name = ws.cell(r, 4).value
            if not name or str(name).strip() == "":
                continue
            changes[str(name).strip()] = {
                "变动情况": str(ws.cell(r, 3).value or "").strip(),
                "变动面积(亩)": str(ws.cell(r, 11).value or "").strip(),
                "变动原因": str(ws.cell(r, 13).value or "").strip(),
                "变动后东至": str(ws.cell(r, 7).value or "").strip(),
                "变动后西至": str(ws.cell(r, 8).value or "").strip(),
                "变动后南至": str(ws.cell(r, 9).value or "").strip(),
                "变动后北至": str(ws.cell(r, 10).value or "").strip(),
            }
    wb.close()

    # Merge: compute actual areas based on changes
    def _safe_float(s):
        try:
            return float(str(s).strip())
        except (ValueError, TypeError):
            return 0.0

    rows = []
    total_actual = 0.0
    for i, p in enumerate(plots, 1):
        ch = changes.pop(p["地块名称"], None)
        confirmed_area = _safe_float(p["地块面积(亩)"])
        if ch:
            change_area = _safe_float(ch["变动面积(亩)"])
            change_type = ch["变动情况"]
            if "增加" in change_type:
                actual_area = confirmed_area + change_area
            elif "减少" in change_type:
                actual_area = confirmed_area - change_area
            else:
                actual_area = confirmed_area
            row = {
                "地块序号": i,
                "地块名称": p["地块名称"],
                "地块编码": p["地块编码"],
                "地块面积(亩)": str(round(actual_area, 4)),
                "东至": ch.get("变动后东至", "") or p["东至"],
                "西至": ch.get("变动后西至", "") or p["西至"],
                "南至": ch.get("变动后南至", "") or p["南至"],
                "北至": ch.get("变动后北至", "") or p["北至"],
                "变动情况": ch["变动情况"],
                "变动面积(亩)": ch["变动面积(亩)"],
                "变动原因": ch["变动原因"],
            }
        else:
            actual_area = confirmed_area
            row = {
                "地块序号": i,
                "地块名称": p["地块名称"],
                "地块编码": p["地块编码"],
                "地块面积(亩)": p["地块面积(亩)"],
                "东至": p["东至"], "西至": p["西至"],
                "南至": p["南至"], "北至": p["北至"],
                "变动情况": "无",
                "变动面积(亩)": "",
                "变动原因": "",
                "变动后东至": "无", "变动后西至": "无",
                "变动后南至": "无", "变动后北至": "无",
            }
        total_actual += actual_area
        rows.append(row)

    # Newly added plots from changes (not in confirmed section)
    for ch in changes.values():
        change_area = _safe_float(ch["变动面积(亩)"])
        total_actual += change_area
        rows.append({
            "地块序号": len(rows) + 1,
            "地块名称": "",
            "地块编码": "",
            "地块面积(亩)": ch["变动面积(亩)"],
            "东至": "", "西至": "",
            "南至": "", "北至": "",
            "变动情况": ch["变动情况"],
            "变动面积(亩)": ch["变动面积(亩)"],
            "变动原因": ch["变动原因"],
            "变动后东至": ch.get("变动后东至", "") or "无",
            "变动后西至": ch.get("变动后西至", "") or "无",
            "变动后南至": ch.get("变动后南至", "") or "无",
            "变动后北至": ch.get("变动后北至", "") or "无",
        })

    # 用计算后的实际总面积替换原始确权总面积
    info["确权总面积(亩)"] = str(round(total_actual, 4))

    return info, rows


def _household_key(row):
    return (row[_HOUSEHOLD_KEY_COLS[0]], row[_HOUSEHOLD_KEY_COLS[1]])


# ── 扫描 ──────────────────────────────────────────────────────────────────────

def scan_folder(folder_path):
    results = []
    errors = []
    xlsx_files = []
    for root, dirs, files in os.walk(folder_path):
        for f in files:
            if _is_biao1(f):
                rel = os.path.relpath(root, folder_path)
                group = rel if rel != "." else os.path.basename(folder_path)
                xlsx_files.append((os.path.join(root, f), group))
    xlsx_files.sort(key=lambda x: x[0])
    total = len(xlsx_files)

    # 预扫描：收集所有已有编码，用于分户时分配新编码
    all_codes = set()
    for fp, _ in xlsx_files:
        try:
            fn = os.path.basename(fp)
            code = fn.split("-")[0] if "-" in fn else ""
            if code.isdigit():
                all_codes.add(int(code))
        except Exception:
            pass
    next_code = max(all_codes) + 1 if all_codes else 1

    # 第一遍：收集确权人员和减少人员，用于跨文件匹配
    confirmed_people = {}  # person_key -> set of file_codes
    reduced_people = {}  # person_key -> set of file_codes (where they were reduced from)
    name_to_pks = {}  # name -> [person_key, ...]
    for fp, _ in xlsx_files:
        try:
            fn = os.path.basename(fp)
            file_code = fn.split("-")[0] if "-" in fn else ""
            groups = parse_biao1(fp, "_")
            for info, rows in groups:
                for row in rows:
                    pk = _person_key(row)
                    name = row.get("家庭成员姓名", "").strip()
                    if name:
                        pks = name_to_pks.setdefault(name, [])
                        if pk not in pks:
                            pks.append(pk)
                    if "减少" in row.get("变动情况", ""):
                        reduced_people.setdefault(pk, set()).add(file_code)
                    else:
                        confirmed_people.setdefault(pk, set()).add(file_code)
        except Exception:
            pass

    for idx, (fp, group) in enumerate(xlsx_files):
        try:
            groups = parse_biao1(fp, group)
            for info, rows in groups:
                if "_split_group" in info:
                    info["承包方编码"] = str(next_code)
                    next_code += 1
                reduce_count = sum(1 for r in rows if "减少" in r.get("变动情况", ""))
                population = len(rows) - reduce_count
                # 查找此户的父户编码（分户）或合户来源
                this_file_code = info.get("编号", "")
                parent_code = ""
                merge_source = ""
                # 工具分户：直接使用 _split_source
                if "_split_group" in info:
                    parent_code = info.get("_split_source", "")
                elif "_split_group" not in info:
                    # 分户检测：确权成员在其他文件中被减少
                    for row in rows:
                        if "减少" not in row.get("变动情况", ""):
                            pk = _person_key(row)
                            parents = reduced_people.get(pk, set())
                            if not parents:
                                name = row.get("家庭成员姓名", "").strip()
                                for cpk in name_to_pks.get(name, []):
                                    if cpk in reduced_people:
                                        parents = parents | reduced_people[cpk]
                            parents = sorted(c for c in parents if c != this_file_code)
                            if parents:
                                parent_code = parents[0]
                                break
                    # 合户检测：变动区“增加”人员在其他文件中是确权成员
                    if not parent_code:
                        for row in rows:
                            if "增加" in row.get("变动情况", ""):
                                pk = _person_key(row)
                                sources = confirmed_people.get(pk, set())
                                if not sources:
                                    name = row.get("家庭成员姓名", "").strip()
                                    for cpk in name_to_pks.get(name, []):
                                        if cpk in confirmed_people:
                                            sources = sources | confirmed_people[cpk]
                                sources = sorted(c for c in sources if c != this_file_code)
                                if sources:
                                    merge_source = sources[0]
                                    break
                for row in rows:
                    merged = {**info, **row}
                    merged["户内人口"] = population
                    code = parent_code or merge_source
                    if (not "减少" in row.get("变动情况", "")) and code:
                        merged["分、合户来源"] = code
                    merged["身份证+性别核对"] = _check_gender(
                        merged.get("身份证号"), merged.get("性别"), merged.get("与承包方代表关系"))
                    c = tuple(merged.get(col, "") for col in OUTPUT_COLUMNS)
                    results.append({"values": c, "key": _household_key(c)})
        except Exception as e:
            errors.append("%s: %s" % (os.path.basename(fp), str(e)))
        yield idx + 1, total, results, errors
def scan_folder_b2(folder_path):
    results = []
    errors = []
    xlsx_files = []
    for root, dirs, files in os.walk(folder_path):
        for f in files:
            if _is_biao2(f):
                rel = os.path.relpath(root, folder_path)
                group = rel if rel != "." else os.path.basename(folder_path)
                xlsx_files.append((os.path.join(root, f), group))
    xlsx_files.sort(key=lambda x: x[0])
    total = len(xlsx_files)

    for idx, (fp, group) in enumerate(xlsx_files):
        try:
            info, rows = parse_biao2(fp, group)
            for row in rows:
                merged = {**info, **row}
                c = tuple(merged.get(col, "") for col in OUTPUT_COLUMNS_B2)
                results.append({"values": c, "key": _household_key(c)})
        except Exception as e:
            errors.append("%s: %s" % (os.path.basename(fp), str(e)))
        yield idx + 1, total, results, errors


# ── GUI ───────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    _tag_colors = {"h0": "#DAEEF3", "h1": "#E2EFDA", "sep": "#F2F2F2"}

    def __init__(self):
        super().__init__()
        self.title("承包方家庭成员 & 承包地块提取工具 v0.3.4")
        self.geometry("1200x650")
        self.minsize(900, 500)
        self.results_b1 = []
        self.results_b2 = []
        self._scanning = False
        self._build_ui()

    def _build_ui(self):
        style = ttk.Style(self)
        style.configure("TButton", padding=(10, 4))
        style.configure("TLabel", padding=(10, 4))
        style.configure("Big.TButton", padding=(10, 6))

        # 文件夹
        top = ttk.Frame(self, padding=(10, 8, 10, 4))
        top.pack(fill=tk.X)
        ttk.Label(top, text="源文件夹：").pack(side=tk.LEFT)
        self.folder_var = tk.StringVar(value=r"D:\农经全延保")
        ttk.Entry(top, textvariable=self.folder_var, width=60).pack(side=tk.LEFT, padx=(0, 6), fill=tk.X, expand=True)
        ttk.Button(top, text="浏览…", command=self._browse).pack(side=tk.LEFT)

        # 按钮行
        btn_frame = ttk.Frame(self, padding=(10, 4))
        btn_frame.pack(fill=tk.X)
        self.start_btn = ttk.Button(btn_frame, text="提取", style="Big.TButton", command=self._start)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 12))
        self.export_btn = ttk.Button(btn_frame, text="导出 Excel", command=self._export, state=tk.DISABLED)
        self.export_btn.pack(side=tk.LEFT, padx=(0, 12))
        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(btn_frame, textvariable=self.status_var).pack(side=tk.LEFT, padx=(12, 0))

        # 进度条
        prog_frame = ttk.Frame(self, padding=(10, 2, 10, 4))
        prog_frame.pack(fill=tk.X)
        self.progress = ttk.Progressbar(prog_frame, mode="determinate", length=400)
        self.progress.pack(fill=tk.X, expand=True)

        # Notebook
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(4, 0))

        # Tab 1
        tab1 = ttk.Frame(self.notebook)
        self.notebook.add(tab1, text="  家庭成员  ")
        self.tree_b1 = ttk.Treeview(tab1, columns=list(range(len(OUTPUT_COLUMNS))), show="headings", height=16)
        self._setup_tree(self.tree_b1, OUTPUT_COLUMNS)
        vsb1 = ttk.Scrollbar(tab1, orient=tk.VERTICAL, command=self.tree_b1.yview)
        hsb1 = ttk.Scrollbar(tab1, orient=tk.HORIZONTAL, command=self.tree_b1.xview)
        self.tree_b1.configure(yscrollcommand=vsb1.set, xscrollcommand=hsb1.set)
        self.tree_b1.grid(row=0, column=0, sticky="nsew")
        vsb1.grid(row=0, column=1, sticky="ns")
        hsb1.grid(row=1, column=0, sticky="ew")
        tab1.rowconfigure(0, weight=1)
        tab1.columnconfigure(0, weight=1)

        # Tab 2
        tab2 = ttk.Frame(self.notebook)
        self.notebook.add(tab2, text="  承包地块  ")
        self.tree_b2 = ttk.Treeview(tab2, columns=list(range(len(OUTPUT_COLUMNS_B2))), show="headings", height=16)
        self._setup_tree(self.tree_b2, OUTPUT_COLUMNS_B2)
        vsb2 = ttk.Scrollbar(tab2, orient=tk.VERTICAL, command=self.tree_b2.yview)
        hsb2 = ttk.Scrollbar(tab2, orient=tk.HORIZONTAL, command=self.tree_b2.xview)
        self.tree_b2.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
        self.tree_b2.grid(row=0, column=0, sticky="nsew")
        vsb2.grid(row=0, column=1, sticky="ns")
        hsb2.grid(row=1, column=0, sticky="ew")
        tab2.rowconfigure(0, weight=1)
        tab2.columnconfigure(0, weight=1)

        # 底部
        bottom = ttk.Frame(self, padding=(10, 2, 10, 8))
        bottom.pack(fill=tk.X)
        self.count_var = tk.StringVar(value="")
        ttk.Label(bottom, textvariable=self.count_var, font=("", 10)).pack(side=tk.LEFT)

    def _setup_tree(self, tree, columns):
        col_widths = {
            "所属组": 60, "承包方编码": 200, "承包合同编号": 200, "承包方代表": 90,
            "发包方名称": 160, "户内人口": 50,
            "家庭成员姓名": 90, "性别": 50, "身份证号": 180, "身份证+性别核对": 60,
            "联系电话": 120,
            "与承包方代表关系": 120, "变动情况": 70,
            "调查记事(附记)": 200,
            "联系方式": 110, "确权总面积(亩)": 90,
            "地块总数": 60, "地块序号": 60,
            "地块名称": 90, "地块编码": 180,
            "地块面积(亩)": 90,
            "东至": 150, "西至": 150, "南至": 150, "北至": 150,
            "变动面积(亩)": 90, "变动原因": 180,
            "调查记事": 200,
        }
        for i, name in enumerate(columns):
            tree.heading(i, text=name)
            tree.column(i, width=col_widths.get(name, 80), minwidth=40)
        for tag, color in self._tag_colors.items():
            tree.tag_configure(tag, background=color)

    def _browse(self):
        d = filedialog.askdirectory(initialdir=self.folder_var.get())
        if d:
            self.folder_var.set(d)

    def _lock(self):
        self.start_btn.configure(state=tk.DISABLED)
        self.export_btn.configure(state=tk.DISABLED)

    def _unlock(self):
        self.start_btn.configure(state=tk.NORMAL)
        if self.results_b1 or self.results_b2:
            self.export_btn.configure(state=tk.NORMAL)

    # ── 统一提取 ──

    def _start(self):
        folder = self.folder_var.get().strip()
        if not os.path.isdir(folder):
            messagebox.showerror('错误', '请选择有效的文件夹。')
            return
        if self._scanning:
            return
        self._scanning = True
        self._lock()
        self.tree_b1.delete(*self.tree_b1.get_children())
        self.tree_b2.delete(*self.tree_b2.get_children())
        self.results_b1 = []
        self.results_b2 = []
        self._all_errors = []
        self.progress['value'] = 0
        self.status_var.set('正在提取…')
        threading.Thread(target=self._run, args=(folder,), daemon=True).start()

    def _run(self, folder):
        # 表1
        results = []
        errors = []
        done = 0
        total = 0
        for done, total, results, errors in scan_folder(folder):
            self.after(0, self._update_progress, done, total * 2, '家庭成员')
        self.results_b1 = results
        self._all_errors.extend(errors)
        b1_total = done if done else 1
        # 表2
        results = []
        errors = []
        done = 0
        for done, total, results, errors in scan_folder_b2(folder):
            self.after(0, self._update_progress, b1_total + done, b1_total + total, '承包地块')
        self.results_b2 = results
        self._all_errors.extend(errors)
        self.after(0, self._finish)

    def _finish(self):
        self._scanning = False
        self._fill_tree(self.tree_b1, self.results_b1, OUTPUT_COLUMNS)
        self._fill_tree(self.tree_b2, self.results_b2, OUTPUT_COLUMNS_B2, is_b2=True)
        self._unlock()
        hk1 = len(set(i['key'] for i in self.results_b1))
        hk2 = len(set(i['key'] for i in self.results_b2))
        self.status_var.set('提取完成')
        self.count_var.set(
            '家庭成员：%d 条(%d 户) | 承包地块：%d 条(%d 户)'
            % (len(self.results_b1), hk1, len(self.results_b2), hk2))
        msg = '提取完成！'
        msg += '\n家庭成员：%d 条(%d 户)' % (len(self.results_b1), hk1)
        msg += '\n承包地块：%d 条(%d 户)' % (len(self.results_b2), hk2)
        if self._all_errors:
            msg += '\n\n错误 %d 条' % len(self._all_errors)
        messagebox.showinfo('完成', msg)


    # ── 公用 ──

    def _update_progress(self, done, total, label):
        self.progress["maximum"] = total
        self.progress["value"] = done
        self.status_var.set("正在提取%s %d/%d …" % (label, done, total))

    # b2 household info columns that should only show once per group
    _B2_HOUSEHOLD_COLS = frozenset(range(7))  # cols 0-5: 所属组, 承包方编码, 承包方代表, 联系方式, 确权总面积, 地块总数
    _B1_HOUSEHOLD_COLS = frozenset({5, 13})  # cols: 户内人口, 调查记事(附记)

    def _fill_tree(self, tree, results, columns, is_b2=False):
        tree.delete(*tree.get_children())
        if not results:
            return
        prev_key = None
        parity = 0
        for item in results:
            cur_key = item["key"]
            if prev_key is not None and cur_key != prev_key:
                parity = 1 - parity
                sep = [""] * len(columns)
                sep[0] = "───"
                tree.insert("", tk.END, values=sep, tags=("sep",))
            tag = "h%d" % parity
            vals = list(item["values"])
            if is_b2 and prev_key == cur_key:
                for ci in self._B2_HOUSEHOLD_COLS:
                    vals[ci] = ""
            if (not is_b2) and prev_key == cur_key:
                for ci in self._B1_HOUSEHOLD_COLS:
                    vals[ci] = ""
            tree.insert("", tk.END, values=vals, tags=(tag,))
            prev_key = cur_key

    def _export(self):
        has_b1 = bool(self.results_b1)
        has_b2 = bool(self.results_b2)
        if not has_b1 and not has_b2:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="汇总表.xlsx",
            initialdir=self.folder_var.get(),
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            if has_b1:
                ws1 = wb.active
                self._export_sheet(ws1, self.results_b1, list(OUTPUT_COLUMNS), "家庭成员")
            if has_b2:
                ws2 = wb.create_sheet() if has_b1 else wb.active
                self._export_sheet(ws2, self.results_b2, list(OUTPUT_COLUMNS_B2), "承包地块", is_b2=True)
            wb.save(path)
            messagebox.showinfo("导出成功", "已保存到：\n%s" % path)
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _export_sheet(self, ws, results, cols, title, is_b2=False):
        ws.title = title
        center_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        thick_bottom = Border(bottom=Side(style="medium", color="808080"))

        for ci, cn in enumerate(cols, 1):
            cell = ws.cell(1, ci, cn)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        cur_row = 2
        prev_key = None
        parity = 0
        for item in results:
            vals = list(item["values"])
            cur_key = item["key"]
            if is_b2 and prev_key == cur_key:
                for ci in range(6):
                    vals[ci] = ""
            if (not is_b2) and prev_key == cur_key:
                for ci in (5, 13):  # 户内人口, 调查记事
                    vals[ci] = ""
            if prev_key is not None and cur_key != prev_key:
                parity = 1 - parity
                for c in range(1, len(cols) + 1):
                    ws.cell(cur_row - 1, c).border = thick_bottom
                for c in range(1, len(cols) + 1):
                    ws.cell(cur_row, c, "").fill = _SEP_FILL
                ws.row_dimensions[cur_row].height = 6
                cur_row += 1
            fill = _FILL_A if parity == 0 else _FILL_B
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(cur_row, ci, val)
                if ci == 10 and val == "错误":
                    cell.fill = _ERR_FILL
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    cell.fill = fill
                if ci <= 8 or ci == 10:
                    cell.alignment = center_align
            prev_key = cur_key
            cur_row += 1

        for c in range(1, len(cols) + 1):
            ws.cell(cur_row - 1, c).border = thick_bottom

        for ci, cn in enumerate(cols, 1):
            max_len = len(cn)
            for r in range(2, min(cur_row, 500)):
                cv = str(ws.cell(r, ci).value or "")
                cl = sum(2 if ord(ch) > 127 else 1 for ch in cv)
                max_len = max(max_len, cl)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 3, 40)
        ws.freeze_panes = "A2"


if __name__ == "__main__":
    app = App()
    app.mainloop()

